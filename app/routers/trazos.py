"""
Router de trazos / placas (Fase A — fases de tela F1–F3), alineado al Excel.
Aditivo: rutas bajo /trazos, no toca el motor de corte por pieza.
"""
import io
from typing import List, Optional
from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session

from app.database.connection import get_db
from app.models.of import OrdenFabricacion
from app.models.usuario import Usuario
from app.core.auth import get_current_user, get_rol
from app.core.templates import templates
from app.core.websocket_manager import ws_manager
from app.models.trazo import OFTrazo, OFTrazoMovimiento
from app.services import trazo_service

router = APIRouter()

from app.roles import ROLES_TRAZO


def _check(user: Usuario):
    if get_rol(user) not in ROLES_TRAZO:
        raise HTTPException(403, f"Rol '{get_rol(user)}' no puede gestionar placas")


class TallaDibujo(BaseModel):
    sku_id: int
    veces: int


class CrearPlacaReq(BaseModel):
    nombre: Optional[str] = None
    largo: Optional[float] = None
    capas: int
    tallas: List[TallaDibujo] = []


class TendidoReq(BaseModel):
    capas: Optional[int] = None


class CorteReq(BaseModel):
    capas: Optional[int] = None


class MaxCapasReq(BaseModel):
    max_capas: Optional[int] = None


def _trazo_dict(tz) -> dict:
    planeado = tz.capas or 0
    tendidas = tz.capas_tendidas or 0
    return {
        "id": tz.id, "nombre": tz.nombre, "largo": tz.largo, "capas": tz.capas,
        "capas_tendidas": tendidas, "capas_restantes": max(0, planeado - tendidas),
        "capas_cortadas": (tz.capas_cortadas or 0), "capas_restantes_corte": max(0, planeado - (tz.capas_cortadas or 0)),
        "metraje": tz.metraje, "total_prendas": tz.total_prendas,
        "estado_tizado": tz.estado_tizado, "estado_tendido": tz.estado_tendido, "estado_corte": tz.estado_corte,
        "tallas": [{"sku_id": t.sku_id, "talla": t.talla, "veces": t.veces, "cantidad": t.cantidad} for t in tz.tallas],
    }


@router.get("/{of_id}", response_class=HTMLResponse)
def armar_placas_page(of_id: int, request: Request, db: Session = Depends(get_db),
                      current_user: Usuario = Depends(get_current_user)):
    of = db.query(OrdenFabricacion).filter_by(id=of_id).first()
    if not of:
        raise HTTPException(404, "OF no encontrada")
    return templates.TemplateResponse("of/trazos.html", {
        "request": request, "of": of,
        "current_user": current_user, "rol": get_rol(current_user),
        "puede_editar": get_rol(current_user) in ROLES_TRAZO,
    })


@router.get("/api/{of_id}/data")
def data(of_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    of = db.query(OrdenFabricacion).filter_by(id=of_id).first()
    if not of:
        raise HTTPException(404, "OF no encontrada")
    metas = trazo_service.meta_por_sku(of_id, db)
    return {
        "of": {"id": of.id, "numero_of": of.numero_of, "total_juegos": of.total_juegos},
        "max_capas": trazo_service.max_capas_of(of),
        "tallas_curva": [
            {"sku_id": sid, "talla": info["talla"], "meta": info["meta"]}
            for sid, info in sorted(metas.items(), key=lambda kv: kv[1]["talla"])
        ],
        "trazos": [_trazo_dict(t) for t in trazo_service.listar_trazos(of_id, db)],
        "validacion": trazo_service.validar_cobertura(of_id, db),
        "consumo": trazo_service.resumen_consumo(of_id, db),
        "fases_tela": trazo_service.fases_tela_info(of_id, db),
    }


@router.post("/api/{of_id}/fase/{fase_tela}/iniciar")
def iniciar_fase(of_id: int, fase_tela: str, db: Session = Depends(get_db),
                 current_user: Usuario = Depends(get_current_user)):
    _check(current_user)
    trazo_service.iniciar_fase_tela(of_id, fase_tela, db)
    ws_manager.notify_of(of_id, "fase_tela", {"fase": fase_tela, "por": current_user.nombre})
    return {"ok": True}


@router.post("/api/{of_id}/crear")
def crear(of_id: int, body: CrearPlacaReq, db: Session = Depends(get_db),
          current_user: Usuario = Depends(get_current_user)):
    _check(current_user)
    tz = trazo_service.crear_trazo(of_id, body.nombre, body.largo, body.capas,
                                   [t.model_dump() for t in body.tallas], db)
    ws_manager.notify_of(of_id, "tela", {"accion": "placa", "por": current_user.nombre})
    return _trazo_dict(tz)


@router.post("/api/{of_id}/max-capas")
def max_capas(of_id: int, body: MaxCapasReq, db: Session = Depends(get_db),
              current_user: Usuario = Depends(get_current_user)):
    _check(current_user)
    mc = trazo_service.set_max_capas(of_id, body.max_capas, db)
    ws_manager.notify_of(of_id, "tela", {"accion": "tope", "por": current_user.nombre})
    return {"max_capas": mc}


@router.delete("/api/trazo/{trazo_id}")
def eliminar(trazo_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    _check(current_user)
    tz = db.query(OFTrazo).filter_by(id=trazo_id).first()
    of_id = tz.of_id if tz else None
    trazo_service.eliminar_trazo(trazo_id, db)
    if of_id:
        ws_manager.notify_of(of_id, "tela", {"accion": "eliminar", "por": current_user.nombre})
    return {"ok": True}


@router.post("/api/trazo/{trazo_id}/tendido")
def tendido(trazo_id: int, body: TendidoReq, db: Session = Depends(get_db),
            current_user: Usuario = Depends(get_current_user)):
    _check(current_user)
    tz = trazo_service.registrar_tendido(trazo_id, body.capas, db, usuario_id=current_user.id)
    ws_manager.notify_of(tz.of_id, "tela", {"accion": "tendido", "por": current_user.nombre})
    return _trazo_dict(tz)


@router.post("/api/trazo/{trazo_id}/corte")
def corte(trazo_id: int, body: CorteReq = CorteReq(), db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    _check(current_user)
    tz = trazo_service.marcar_corte(trazo_id, body.capas, db, usuario_id=current_user.id)
    ws_manager.notify_of(tz.of_id, "tela", {"accion": "corte", "por": current_user.nombre})
    return _trazo_dict(tz)


@router.get("/api/{of_id}/reporte-excel")
def reporte_excel(of_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Exporta placas, consumo de tela e historial de tendido/corte de la OF a Excel."""
    of = db.query(OrdenFabricacion).filter_by(id=of_id).first()
    if not of:
        raise HTTPException(404, "OF no encontrada")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl no instalado. Ejecuta: pip install openpyxl")

    hdr_fill = PatternFill("solid", fgColor="1A5FA0")
    hdr_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    def _encabezar(ws, headers):
        ws.append(headers)
        for c in ws[ws.max_row]:
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")

    wb = Workbook()

    # Hoja 1: Placas
    ws = wb.active; ws.title = "Placas"
    ws.append([f"OF {of.numero_of}", of.cliente]); ws["A1"].font = bold
    ws.append([f"Prenda: {of.prenda_catalogo.codigo if of.prenda_catalogo else of.tipo_prenda}",
               f"Total prendas: {of.total_juegos}"])
    ws.append([])
    _encabezar(ws, ["Placa", "Capas", "Tendidas", "Cortadas", "Largo (m)", "Metros", "Tizado", "Tendido", "Corte", "Prendas"])
    tot_metros = 0.0; tot_prendas = 0
    for t in trazo_service.listar_trazos(of_id, db):
        ws.append([t.nombre, t.capas, t.capas_tendidas or 0, t.capas_cortadas or 0,
                   t.largo, t.metraje, t.estado_tizado, t.estado_tendido, t.estado_corte, t.total_prendas])
        tot_metros += (t.metraje or 0); tot_prendas += (t.total_prendas or 0)
    fila_tot = ["TOTAL", "", "", "", "", round(tot_metros, 1), "", "", "", tot_prendas]
    ws.append(fila_tot)
    for c in ws[ws.max_row]:
        c.font = bold
    for col, w in zip("ABCDEFGHIJ", [16, 8, 9, 9, 9, 9, 10, 10, 10, 9]):
        ws.column_dimensions[col].width = w

    # Hoja 2: Consumo
    cons = trazo_service.resumen_consumo(of_id, db)
    ws2 = wb.create_sheet("Consumo")
    _encabezar(ws2, ["Métrica", "Valor"])
    ws2.append(["Proyectado (HDC) m/prenda", cons.get("proyectado")])
    ws2.append(["Real ponderado m/prenda", cons.get("real")])
    ws2.append(["Metros totales", cons.get("metros")])
    ws2.append(["Prendas", cons.get("prendas")])
    ws2.append(["Desvío (m)", cons.get("desvio")])
    ws2.append(["Desvío (%)", cons.get("desvio_pct")])
    ws2.column_dimensions["A"].width = 28; ws2.column_dimensions["B"].width = 14

    # Hoja 3: Historial tendido/corte
    ws3 = wb.create_sheet("Historial tela")
    _encabezar(ws3, ["Fecha", "Placa", "Tipo", "Capas", "Acumulado", "Usuario"])
    movs = (
        db.query(OFTrazoMovimiento, OFTrazo.nombre)
        .join(OFTrazo, OFTrazo.id == OFTrazoMovimiento.trazo_id)
        .filter(OFTrazo.of_id == of_id)
        .order_by(OFTrazoMovimiento.created_at.desc())
        .all()
    )
    for mov, placa in movs:
        ws3.append([
            mov.created_at.strftime("%d/%m/%Y %H:%M") if mov.created_at else "",
            placa, mov.tipo, mov.capas, mov.acumulado,
            mov.usuario.nombre if mov.usuario else "",
        ])
    for col, w in zip("ABCDEF", [17, 14, 10, 8, 10, 22]):
        ws3.column_dimensions[col].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"OF_{of.numero_of}_placas.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/api/{of_id}/movimientos")
def movimientos(of_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Historial por sesión de tendido/corte de todas las placas de la OF."""
    q = (
        db.query(OFTrazoMovimiento, OFTrazo.nombre)
        .join(OFTrazo, OFTrazo.id == OFTrazoMovimiento.trazo_id)
        .filter(OFTrazo.of_id == of_id)
        .order_by(OFTrazoMovimiento.created_at.desc())
        .limit(200)
    )
    return [
        {
            "placa": placa, "tipo": mov.tipo, "capas": mov.capas, "acumulado": mov.acumulado,
            "usuario": mov.usuario.nombre if mov.usuario else None,
            "fecha": mov.created_at.strftime("%d/%m %H:%M") if mov.created_at else None,
        }
        for mov, placa in q.all()
    ]
