"""
Lee la Hoja de Costos (HDC) real de Schellenger y carga materiales (MP) + avíos
COMPLETOS (con precio, moneda, factor, proveedor, consumo) en las bases:
  hoja 'MODERN FIT' -> base SCH-MF   |   hoja 'SLIM FIT' -> base SCH-SF

Reemplaza la carga hardcodeada: toma los valores exactos del HDC.
Recarga limpio (borra MP/avíos previos de la base) para incluir los precios.

Uso:  python seed_hdc_schellenger.py "ruta\\HDC-CAMISAS - SCHELLENGER MF y SF.xlsx"
"""
import sys
import openpyxl
from app.database.connection import SessionLocal
import app.models.usuario, app.models.of, app.models.pieza, app.models.fase       # noqa: F401
import app.models.planta, app.models.catalogo, app.models.curva_tallas            # noqa: F401
import app.models.ingenieria, app.models.parametro, app.models.trazo, app.models.paquete  # noqa: F401
from app.models.catalogo import (PrendaCatalogo, CatalogoMp, CatalogoAvio,
                                  PrendaMpConfig, PrendaAvioConfig,
                                  CatalogoServicio, CatalogoMod, SERVICIOS_TERCEROS)

HEADER_ROW = 36                      # fila de encabezados del HDC
SECCIONES = {"CORTE", "COSTURA", "ACABADOS", "EMBALAJE", "HABILITADO"}
FOOTER = ("ELABORADO", "VºBº", "V°B°", "VB")
OPERACIONES_MOD = {"CORTE", "COSTURA", "ACABADO", "HABILITADO", "BORDADO", "LAVADO"}
# hoja HDC -> codigo de base
SHEETS = {"MODERN FIT": "SCH-MF", "SLIM FIT": "SCH-SF"}

# columnas (1-based) confirmadas del HDC
C_COD, C_INS, C_PROV, C_PROC = 2, 3, 7, 8
C_ANCHO, C_UM, C_CONS, C_ADIC = 10, 11, 12, 13
C_FACTOR, C_UC, C_MON, C_PRECIO = 15, 16, 17, 18
# columnas de MOD / servicios
C_MINSTD, C_EFIC, C_COSTOMIN, C_SUBTOTAL = 16, 18, 23, 24


def _num(v, d=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return d


def _tipo_mp(nombre):
    n = (nombre or "").upper()
    if "ENTRETELA" in n or "REFUERZO" in n:
        return "ENTRETELA"
    if "TELA" in n or "COTTON" in n or "POLYESTER" in n or "ALGODON" in n:
        return "TELA_PRINCIPAL"
    return "ACCESORIO"


def _parse_hoja(ws):
    """Devuelve (materiales, avios, servicios, mods) como listas de dicts."""
    mps, avios, servicios, mods, seccion, bloque = [], [], [], [], None, "MATERIALES"
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        g = lambda c: ws.cell(row=r, column=c).value
        cod = (str(g(C_COD)).strip() if g(C_COD) is not None else "")
        nom = (str(g(C_INS)).strip() if g(C_INS) is not None else "")
        a1 = (str(g(1)).strip().upper() if g(1) is not None else "")
        codu = cod.upper()

        # cambios de bloque
        if "OTROS" in codu and "SERVICIO" in codu:
            bloque = "SERVICIOS"; continue
        if codu.startswith("MANO DE OBRA"):
            bloque = "MOD"; continue
        if codu.startswith(("COSTO", "GASTOS", "TOTAL", "CONFECC")) or codu.startswith(("C.V", "CV")):
            if bloque == "MOD" and codu.startswith("CONFECC"):
                continue                     # encabezado de la tabla MOD, no corta
            bloque = "FIN"
        if bloque == "FIN":
            continue

        if bloque == "SERVICIOS":
            if codu in SERVICIOS_TERCEROS:
                costo = _num(g(C_SUBTOTAL), None) if g(C_SUBTOTAL) is not None else None
                if costo:                    # solo los que aplican (con costo)
                    servicios.append(dict(nombre=codu, costo=costo,
                                          moneda=(str(g(C_MON)).strip() if g(C_MON) else None)))
            continue

        if bloque == "MOD":
            if codu in OPERACIONES_MOD and isinstance(g(C_MINSTD), (int, float)):
                mods.append(dict(operacion=codu, min_std=_num(g(C_MINSTD), 0.0),
                                 efic=_num(g(C_EFIC), 1.0) or 1.0,
                                 costo_min=_num(g(C_COSTOMIN), 0.0)))
            continue

        # bloque MATERIALES
        if codu in SECCIONES:
            seccion = codu; continue
        if not nom:
            continue
        if any(nom.upper().startswith(f) for f in FOOTER):
            continue
        if a1 == "NO":
            continue
        # El HDC costea (col24) solo algunos ítems; si col24≈0, el ítem NO se costea
        # (se carga en la ficha para el BOM pero sin precio) → el total cuadra con el HDC.
        c24 = _num(g(C_SUBTOTAL), None) if g(C_SUBTOTAL) is not None else None
        costeado = c24 is not None and abs(c24) > 1e-6
        precio = (_num(g(C_PRECIO), None) if (costeado and g(C_PRECIO) is not None) else None)
        d = dict(codigo=cod or None, nombre=nom,
                 proveedor=(str(g(C_PROV)).strip() if g(C_PROV) else None),
                 procedencia=(str(g(C_PROC)).strip() if g(C_PROC) else None),
                 unidad_medida=(str(g(C_UM)).strip() if g(C_UM) else "Unid"),
                 consumo=_num(g(C_CONS), 1.0), adic=_num(g(C_ADIC), 0.01),
                 factor=_num(g(C_FACTOR), 1.0) or 1.0,
                 uc=(str(g(C_UC)).strip() if g(C_UC) else None),
                 moneda=(str(g(C_MON)).strip() if g(C_MON) else None),
                 precio=precio,
                 ancho=_num(g(C_ANCHO), None) if g(C_ANCHO) is not None else None)
        if seccion == "CORTE":
            mps.append(d)
        else:
            d["seccion"] = seccion or "ACABADOS"
            avios.append(d)
    return mps, avios, servicios, mods


def _recargar_base(db, base, mps, avios, servicios, mods):
    # limpiar overrides de variantes y ficha previa de la base
    var_ids = [v.id for v in base.variantes]
    ids = var_ids + [base.id]
    if ids:
        db.query(PrendaMpConfig).filter(PrendaMpConfig.prenda_catalogo_id.in_(ids)).delete(synchronize_session=False)
        db.query(PrendaAvioConfig).filter(PrendaAvioConfig.prenda_catalogo_id.in_(ids)).delete(synchronize_session=False)
    db.query(CatalogoMp).filter_by(prenda_catalogo_id=base.id).delete(synchronize_session=False)
    db.query(CatalogoAvio).filter_by(prenda_catalogo_id=base.id).delete(synchronize_session=False)
    db.query(CatalogoServicio).filter_by(prenda_catalogo_id=base.id).delete(synchronize_session=False)
    db.query(CatalogoMod).filter_by(prenda_catalogo_id=base.id).delete(synchronize_session=False)
    db.flush()
    for i, m in enumerate(mps):
        db.add(CatalogoMp(prenda_catalogo_id=base.id, nombre=m["nombre"], tipo=_tipo_mp(m["nombre"]),
                          ancho_referencia=m["ancho"], consumo_unitario=m["consumo"], pct_adicional=m["adic"],
                          unidad_medida=m["unidad_medida"], unidad_compra=m["uc"], factor_conversion=m["factor"],
                          codigo_interno=m["codigo"], proveedor=m["proveedor"], procedencia=m["procedencia"],
                          moneda=m["moneda"], precio_referencia=m["precio"], orden=i, activo=True))
    for i, a in enumerate(avios):
        db.add(CatalogoAvio(prenda_catalogo_id=base.id, seccion=a["seccion"], nombre=a["nombre"],
                            unidad_medida=a["unidad_medida"], consumo_unitario=a["consumo"], pct_adicional=a["adic"],
                            unidad_compra=a["uc"], factor_conversion=a["factor"], codigo_interno=a["codigo"],
                            proveedor=a["proveedor"], procedencia=a["procedencia"], moneda=a["moneda"],
                            precio=a["precio"], orden=i, activo=True))
    for i, s in enumerate(servicios):
        db.add(CatalogoServicio(prenda_catalogo_id=base.id, nombre=s["nombre"], costo=s["costo"],
                                moneda=s["moneda"], orden=i, activo=True))
    for i, m in enumerate(mods):
        db.add(CatalogoMod(prenda_catalogo_id=base.id, operacion=m["operacion"], min_std=m["min_std"],
                           pct_eficiencia=m["efic"], costo_minuto=m["costo_min"], orden=i, activo=True))
    return len(mps), len(avios), len(servicios), len(mods)


def main():
    if len(sys.argv) < 2:
        print('Uso: python seed_hdc_schellenger.py "ruta\\HDC-CAMISAS - SCHELLENGER MF y SF.xlsx"')
        return
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
    # mapear hojas por nombre normalizado
    hojas = {s.strip().upper(): s for s in wb.sheetnames}
    db = SessionLocal()
    try:
        for hoja_norm, base_cod in SHEETS.items():
            real = hojas.get(hoja_norm)
            if not real:
                print(f"  Hoja '{hoja_norm}' no encontrada en el archivo."); continue
            base = db.query(PrendaCatalogo).filter_by(codigo=base_cod).first()
            if not base:
                print(f"  Base {base_cod} no existe (corre seed_schellenger.py)."); continue
            mps, avios, servicios, mods = _parse_hoja(wb[real])
            nm, na, ns, nmod = _recargar_base(db, base, mps, avios, servicios, mods)
            print(f"  {base_cod} ({hoja_norm}): {nm} MP + {na} avíos + {ns} servicios + {nmod} MOD.")
        db.commit()
        print("Listo. Ficha de MP+avíos completa (con precios) en las bases. Variantes heredan.")
    except Exception as e:
        db.rollback(); print("ERROR:", e); raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
